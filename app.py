from flask import Flask, render_template, request, send_file
import openpyxl
from openpyxl.utils import get_column_letter
import io
import os

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

# A4 printable area with minimal margins (cm)
MARGIN_CM     = 0.5
A4_PORTRAIT_W = 21.0 - 2 * MARGIN_CM   # 20.0 cm
A4_PORTRAIT_H = 29.7 - 2 * MARGIN_CM   # 28.7 cm

DPI = 96.0

# Column width: empirical constant derived from Excel default (8.43 units = 64 px)
# This already accounts for font padding — do NOT add a separate padding term.
PX_PER_COL_UNIT = 64.0 / 8.43   # ≈ 7.59 px per column-width unit

# Row height: simple linear conversion (points → pixels)
PX_PER_ROW_PT = DPI / 72.0      # ≈ 1.333 px per point

DEFAULT_COL_WIDTH  = 8.43
DEFAULT_ROW_HEIGHT = 15.0


def build_col_width_map(ws):
    """
    Correctly expands XML <col min=m max=n width=w> range definitions.
    Accessing ws.column_dimensions[letter] directly only returns the range's
    first column; other columns in the range silently get the wrong default.
    """
    widths = {}
    default = ws.sheet_format.defaultColWidth or DEFAULT_COL_WIDTH
    for dim in ws.column_dimensions.values():
        for col_idx in range(dim.min, dim.max + 1):
            widths[col_idx] = dim.width if (dim.width and dim.width > 0) else default
    return widths


def build_row_height_map(ws):
    heights = {}
    for row_num, dim in ws.row_dimensions.items():
        heights[row_num] = dim.height if (dim.height and dim.height > 0) else DEFAULT_ROW_HEIGHT
    return heights


def fit_sheet_to_a4(ws):
    n_cols = ws.max_column or 1
    n_rows = ws.max_row or 1

    col_map = build_col_width_map(ws)
    row_map = build_row_height_map(ws)

    col_widths  = [col_map.get(c, DEFAULT_COL_WIDTH)  for c in range(1, n_cols + 1)]
    row_heights = [row_map.get(r, DEFAULT_ROW_HEIGHT) for r in range(1, n_rows + 1)]

    # セルを A4 より少し大きく設定 → fitToPage が縮小して用紙いっぱいに収まる
    # ちょうど A4 サイズにすると丸め誤差で 2 ページ or 余白残りが起きる
    OVERSHOOT_W = 1.10
    OVERSHOOT_H = 4.00
    target_w_px = A4_PORTRAIT_W / 2.54 * DPI * OVERSHOOT_W
    target_h_px = A4_PORTRAIT_H / 2.54 * DPI * OVERSHOOT_H

    # 開始時間(N-Q=14-17)・終了時間(R-U=18-21) を他の列より広くする
    TIME_COL_BOOST = 1.5
    TIME_COLS = set(range(14, 22))  # N,O,P,Q,R,S,T,U
    boosted_widths = [w * TIME_COL_BOOST if (i + 1) in TIME_COLS else w
                      for i, w in enumerate(col_widths)]

    total_w_px = sum(boosted_widths) * PX_PER_COL_UNIT
    scale_w = target_w_px / total_w_px

    for col_num, w in enumerate(boosted_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = w * scale_w

    # ヘッダー/フッターは元の高さのまま、残り高さをすべてデータ行（12〜42）へ
    target_h_pt = target_h_px / PX_PER_ROW_PT

    HEADER_END = 11
    DATA_START = 12
    DATA_END = min(42, n_rows)

    header_h = [row_heights[r - 1] for r in range(1, HEADER_END + 1) if r <= n_rows]
    data_h   = [row_heights[r - 1] for r in range(DATA_START, DATA_END + 1) if r <= n_rows]
    footer_h = [row_heights[r - 1] for r in range(DATA_END + 1, n_rows + 1)]

    # ヘッダー/フッターは生の高さで固定し、残りをすべてデータ行に割り当てる
    data_scale = (target_h_pt - sum(header_h) - sum(footer_h)) / sum(data_h) if sum(data_h) > 0 else 1.0

    for row_num in range(1, HEADER_END + 1):
        if row_num <= n_rows:
            ws.row_dimensions[row_num].height = row_heights[row_num - 1]

    for row_num in range(DATA_START, DATA_END + 1):
        if row_num <= n_rows:
            ws.row_dimensions[row_num].height = row_heights[row_num - 1] * data_scale

    for row_num in range(DATA_END + 1, n_rows + 1):
        ws.row_dimensions[row_num].height = row_heights[row_num - 1]

    # データ行（12〜42）の入力済みフィールドを消去
    for row in range(12, 43):
        for col in ['N', 'R', 'AE', 'AZ', 'BH']:  # 開始時間, 終了時間, 食事提供加算, 移行準備支援体制加算, 備考
            ws[f'{col}{row}'].value = None

    # 合計回数: 数式を消して「回」だけ残す
    for coord in ['V43', 'Z43', 'AE43', 'AH43', 'AK43', 'AN43', 'AQ43', 'AT43', 'AW43']:
        ws[coord].value = '回'
    # 移行準備支援体制加算: 当月のみ消去（累計 BH44 は残す）
    ws['BH43'].value = None
    # 当月算定日数
    ws['AZ46'].value = None

    # fitToPage で確実に 1 ページ・用紙いっぱいに収める
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1

    margin_inch = MARGIN_CM / 2.54
    ws.page_margins.top    = margin_inch
    ws.page_margins.bottom = margin_inch
    ws.page_margins.left   = margin_inch
    ws.page_margins.right  = margin_inch
    ws.page_margins.header = 0
    ws.page_margins.footer = 0


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return 'ファイルが選択されていません', 400

    file = request.files['file']
    if not file.filename:
        return 'ファイルが選択されていません', 400

    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        return 'Excelファイル (.xlsx) のみ対応しています', 400

    wb = openpyxl.load_workbook(file)

    for ws in wb.worksheets:
        fit_sheet_to_a4(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    base_name = os.path.splitext(file.filename)[0]
    download_name = f"{base_name}_手書き用.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )


if __name__ == '__main__':
    app.run(debug=True, port=5050)
